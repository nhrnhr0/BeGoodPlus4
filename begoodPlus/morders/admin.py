from django.http import HttpRequest
from .models import MorderStatus
from ordered_model.admin import OrderedModelAdmin
from django.utils.html import mark_safe
from django.contrib import messages
from reversion.admin import VersionAdmin
import datetime
from django.contrib import admin
from openpyxl import Workbook
from docsSignature.utils import create_signature_doc_from_morder
from productColor.models import ProductColor
from django.db import models
from django.forms import Textarea

from productSize.models import ProductSize
from .models import MOrderItem, MOrder, MOrderItemEntry, ProviderRequest, TakenInventory
import io
from django.contrib import admin
from django.http.response import HttpResponse
from django.utils.translation import gettext_lazy as _
import zipfile
from openpyxl.styles import PatternFill

from openpyxl.styles import Alignment
from openpyxl.styles import Font
import copy
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation


class TakenInventoryAdmin(admin.ModelAdmin):
    list_display = ('quantity', 'color', 'size', 'varient',
                    'has_physical_barcode', 'provider')
    list_filter = ('color', 'size', 'varient', 'provider')
    search_fields = ('color', 'size', 'varient',
                     'has_physical_barcode', 'provider')
    ordering = ('color', 'size', 'varient', 'provider')
    pass


admin.site.register(TakenInventory, TakenInventoryAdmin)


class MOrderItemEntryAdmin(admin.ModelAdmin):
    list_display = ('id', 'orderItem', 'color', 'size', 'varient', 'quantity')
    list_filter = ('orderItem', 'color', 'size', 'varient')
    search_fields = ('orderItem', 'color', 'size', 'varient')


admin.site.register(MOrderItemEntry, MOrderItemEntryAdmin)

# Register your models here.


class MOrderItemAdmin(admin.ModelAdmin):
    model = MOrderItem
    list_display = ('id', 'product', 'price', 'ergent',
                    'prining', 'embroidery', 'comment',)
    filter_horizontal = ('providers', 'entries', 'morder', 'taken')


admin.site.register(MOrderItem, MOrderItemAdmin)


class ProviderRequestAdmin(admin.ModelAdmin):
    model = ProviderRequest
    list_display = ('id', 'provider', 'size', 'varient', 'color', 'force_physical_barcode',
                    'quantity', 'product_name_display', 'morder_id_display')

    def product_name_display(self, obj):
        item = obj.orderItem.first()
        if item:
            return item.product.title
        return None
    product_name_display.short_description = _('Product Name')

    def morder_id_display(self, obj):
        item = obj.orderItem.first()
        if item:
            return item.morder.first().id
        return None
    morder_id_display.short_description = _('MOrder id')


admin.site.register(ProviderRequest, ProviderRequestAdmin)


class MOrderAdmin(VersionAdmin):  # admin.ModelAdmin
    model = MOrder
    fields = ('cart', 'total_sell_price', 'client', 'name', 'phone', 'email',
              'status', 'status2', 'message',)  # what is this for?
    readonly_fields = ('created', 'total_sell_price', 'updated', 'get_edit_url',
                       'view_morder_pdf_link', 'cart', 'client', 'status', 'status2', )
    list_display = ('id', 'client', 'name', 'status', 'status2', 'status_msg', 'total_sell_price',
                    'get_edit_url', 'view_morder_pdf_link', 'created', 'updated',)
    # list_editable = ('status_msg',)
    # filter_horizontal = ('products',)
    list_filter = ('status', 'created', 'updated',)
    search_fields = ('id', 'name', 'phone', 'email', 'status', 'message', 'products__product__title',
                     'client__businessName', 'client__email', 'client__extraName', 'client__contactMan', 'client__user__username')
    list_select_related = ('client', 'client__user',)
    actions = ('export_to_excel', 'export_to_signiture_doc',
               'sync_with_spreedsheet')
    # select_related = ('client', 'status2')

    def get_queryset(self, request: HttpRequest):
        return super().get_queryset(request).select_related('client', 'status2', 'client__user')

    formfield_overrides = {
        models.TextField: {'widget': Textarea(
            attrs={'rows': 2,
                   'cols': 20, })},  # 'style': 'height: 1em;'
    }

    def sync_with_spreedsheet(self, request, queryset):
        for morder in queryset:
            morder.morder_to_spreedsheet()
            messages.add_message(
                request, messages.INFO, f'הזמנה {morder.id} סונכרנה')

    def export_to_signiture_doc(self, request, queryset):
        for morder in queryset:
            sigObj = create_signature_doc_from_morder(morder)
            if sigObj:
                # /admin/docsSignature/mordersignature/1/change/
                messages.add_message(
                    request, messages.INFO, mark_safe(f'מסמך חתימה {sigObj.id} <a href="/admin/docsSignature/mordersignature/{sigObj.id}/change/"> לחץ כאן </a>'))
        messages.add_message(request, messages.INFO, mark_safe(
            'לדף כל המסמכים <a href="/admin/docsSignature/mordersignature/"> לחץ כאן </a>'))

    def export_to_excel(self, request, queryset):
        filesbuffers = []

        yellow = "00FFFF00"
        grey = "00808080"
        light_blue = "00FFCCFF"
        header_font = Font(size=10, bold=True,)
        center_align = Alignment(horizontal='center', vertical='center')
        align_rtl = Alignment(horizontal='right',
                              vertical='center', wrap_text=True)
        header_fill = PatternFill(
            start_color=light_blue, end_color=light_blue, fill_type='solid')

        bottom_border = Border(bottom=Side(style='thin'))

        orders = []
        for morder in queryset:
            order_data = morder.get_exel_data()
            orders.append(order_data)
        all_products = {}
        for order in orders:
            name = order['name']
            order_products = order['products']
            for order_product in order_products:
                product_name = order_product['title']
                if product_name not in all_products:
                    all_products[product_name] = {
                        'entries': copy.deepcopy(order_product['entries']),
                        'comment': (name + ': ' + order_product['comment'] + '\n') if order_product['comment'] else '',
                        'barcode': order_product['barcode'],
                        'total_quantity': 0,
                        'details': [],
                    }
                else:
                    # all_products[product_name] += product['entries']
                    entries = all_products[product_name]['entries']
                    for entry in order_product['entries'].keys():
                        if entry in entries:
                            entries[entry] += order_product['entries'][entry]
                        else:
                            entries[entry] = order_product['entries'][entry]
                    all_products[product_name]['entries'] = entries

                    all_products[product_name]['comment'] += (
                        name + ': ' + order_product['comment'] + '\n') if order_product['comment'] else ''
                all_products[product_name]['total_quantity'] += order_product['total_quantity']
                all_products[product_name]['details'].append(
                    name + ' - ' + str(order_product['total_quantity']))
            # create total of quantity, total price,   the products by name
            pass
        # current all_products:
        # {'product_name': {('color_name', 'size_name', 'varient_name'): quantity}}
        # create excel file
        # headers:
        # פריט, הערות, ברקוד, כמות כוללת
        # פירוט מידות צבעים
        headers = ['ברקוד', 'פריט', 'כמות כוללת', 'הערות',
                   'כמות נלקחת', 'מחיר מכירה', 'מע"מ', 'הדפסה?', '', 'רקמה?', '', ]

        wb = Workbook()
        main_ws = wb.active
        main_ws.sheet_view.rightToLeft = True
        # text align all the document to center
        # set the width of the columns
        main_ws.column_dimensions['A'].width = 20
        main_ws.column_dimensions['B'].width = 20
        main_ws.column_dimensions['C'].width = 20
        main_ws.column_dimensions['D'].width = 20
        main_ws.column_dimensions['E'].width = 15
        main_ws.column_dimensions['F'].width = 15
        main_ws.column_dimensions['G'].width = 15
        main_ws.column_dimensions['H'].width = 15
        main_ws.column_dimensions['I'].width = 15
        main_ws.column_dimensions['J'].width = 15
        # main_ws.cell(row=1, column=6).value = 'לא כולל'
        # main_ws.cell(row=2, column=6).value = 'כולל'

        # data_val = DataValidation(
        #     type="list", formula1='=Sheet!$F$1:Sheet!$F$2')
        # main_ws.add_data_validation(data_val)
        # data_val.add(main_ws['B1'])
        # set the header
        ws_rows_counter = 1

        # create sheet for each order
        for order in orders:
            name = order['name']
            order_ws = wb.create_sheet(order['name'] + ' ' + str(order['id']))
            # set the height of the row
            # order_ws.add_data_validation(data_val)

            order_ws.sheet_view.rightToLeft = True
            # text align all the document to center
            # set the width of the columns
            order_ws.column_dimensions['A'].width = 20
            order_ws.column_dimensions['B'].width = 20
            order_ws.column_dimensions['C'].width = 20
            order_ws.column_dimensions['D'].width = 20
            order_ws.column_dimensions['E'].width = 15
            order_ws.column_dimensions['F'].width = 15
            order_ws.column_dimensions['G'].width = 15
            order_ws.column_dimensions['H'].width = 15
            order_ws.column_dimensions['I'].width = 15
            order_ws.column_dimensions['J'].width = 15
            order_products = order['products']
            order_ws_rows_counter = 1
            order_ws.cell(row=order_ws_rows_counter,
                          column=1).value = 'מספר הזמנה'
            order_ws.cell(row=order_ws_rows_counter, column=1).font = order_ws.cell(
                row=order_ws_rows_counter, column=1).font.copy(bold=True)
            order_ws.cell(row=order_ws_rows_counter,
                          column=1).alignment = align_rtl
            order_ws.cell(row=order_ws_rows_counter,
                          column=2).value = 'תאריך הזמנה'
            order_ws.cell(row=order_ws_rows_counter, column=2).font = order_ws.cell(
                row=order_ws_rows_counter, column=2).font.copy(bold=True)
            order_ws.cell(row=order_ws_rows_counter,
                          column=2).alignment = align_rtl
            order_ws.cell(row=order_ws_rows_counter,
                          column=3).value = 'שם הלקוח'
            order_ws.cell(row=order_ws_rows_counter, column=3).font = order_ws.cell(
                row=order_ws_rows_counter, column=3).font.copy(bold=True)
            order_ws.cell(row=order_ws_rows_counter,
                          column=3).alignment = align_rtl
            order_ws.cell(row=order_ws_rows_counter, column=4).value = 'הודעה'
            order_ws.cell(row=order_ws_rows_counter, column=4).font = order_ws.cell(
                row=order_ws_rows_counter, column=4).font.copy(bold=True)
            order_ws.cell(row=order_ws_rows_counter,
                          column=4).alignment = align_rtl
            order_ws_rows_counter += 1
            order_ws.cell(row=order_ws_rows_counter,
                          column=1).value = order['id']
            order_ws.cell(row=order_ws_rows_counter,
                          column=1).alignment = align_rtl
            order_ws.cell(row=order_ws_rows_counter,
                          column=2).value = order['date']
            order_ws.cell(row=order_ws_rows_counter,
                          column=2).alignment = align_rtl
            order_ws.cell(row=order_ws_rows_counter,
                          column=3).value = order['name']
            order_ws.cell(row=order_ws_rows_counter,
                          column=3).alignment = align_rtl
            order_ws.cell(row=order_ws_rows_counter,
                          column=4).value = order['message']
            order_ws.cell(row=order_ws_rows_counter,
                          column=4).alignment = align_rtl
            order_ws_rows_counter += 1
            for i in range(len(headers)):
                order_ws.cell(row=order_ws_rows_counter,
                              column=i+1).value = headers[i]
                order_ws.cell(row=order_ws_rows_counter, column=i+1).font = main_ws.cell(
                    row=order_ws_rows_counter, column=i+1).font.copy(bold=True)
            order_ws_rows_counter += 1
            for product in order_products:
                product_name = product['title']
                order_ws.cell(row=order_ws_rows_counter,
                              column=1).value = product['barcode']
                order_ws.cell(row=order_ws_rows_counter,
                              column=1).fill = header_fill
                order_ws.cell(row=order_ws_rows_counter,
                              column=1).alignment = align_rtl
                order_ws.cell(row=order_ws_rows_counter,
                              column=1).font = header_font
                order_ws.cell(row=order_ws_rows_counter,
                              column=1).border = bottom_border

                order_ws.cell(row=order_ws_rows_counter,
                              column=2).value = product_name
                order_ws.cell(row=order_ws_rows_counter,
                              column=2).fill = header_fill
                order_ws.cell(row=order_ws_rows_counter,
                              column=2).alignment = align_rtl
                order_ws.cell(row=order_ws_rows_counter,
                              column=2).font = header_font
                order_ws.cell(row=order_ws_rows_counter,
                              column=2).border = bottom_border

                order_ws.cell(row=order_ws_rows_counter,
                              column=3).value = product['total_quantity']
                order_ws.cell(row=order_ws_rows_counter,
                              column=3).fill = header_fill
                order_ws.cell(row=order_ws_rows_counter,
                              column=3).alignment = align_rtl
                order_ws.cell(row=order_ws_rows_counter,
                              column=3).font = header_font
                order_ws.cell(row=order_ws_rows_counter,
                              column=3).border = bottom_border

                order_ws.cell(row=order_ws_rows_counter,
                              column=4).value = product['comment']
                order_ws.cell(row=order_ws_rows_counter,
                              column=4).fill = header_fill
                order_ws.cell(row=order_ws_rows_counter,
                              column=4).alignment = align_rtl
                order_ws.cell(row=order_ws_rows_counter,
                              column=4).font = header_font
                order_ws.cell(row=order_ws_rows_counter,
                              column=4).border = bottom_border

                order_ws.cell(row=order_ws_rows_counter,
                              column=5).value = ''  # taken
                order_ws.cell(row=order_ws_rows_counter,
                              column=5).fill = header_fill
                order_ws.cell(row=order_ws_rows_counter,
                              column=5).alignment = align_rtl
                order_ws.cell(row=order_ws_rows_counter,
                              column=5).font = header_font
                order_ws.cell(row=order_ws_rows_counter,
                              column=5).border = bottom_border

                order_ws.cell(row=order_ws_rows_counter, column=6).value = str(
                    product['price']) + '₪'
                order_ws.cell(row=order_ws_rows_counter,
                              column=6).fill = header_fill
                order_ws.cell(row=order_ws_rows_counter,
                              column=6).alignment = align_rtl
                order_ws.cell(row=order_ws_rows_counter,
                              column=6).font = header_font
                order_ws.cell(row=order_ws_rows_counter,
                              column=6).border = bottom_border

                order_ws.cell(row=order_ws_rows_counter,
                              column=7).value = 'לא כולל'
                order_ws.cell(row=order_ws_rows_counter,
                              column=7).fill = header_fill
                order_ws.cell(row=order_ws_rows_counter,
                              column=7).alignment = align_rtl
                order_ws.cell(row=order_ws_rows_counter,
                              column=7).font = header_font
                order_ws.cell(row=order_ws_rows_counter,
                              column=7).border = bottom_border
                # data_val.add(order_ws.cell(
                #     row=order_ws_rows_counter, column=7))

                order_ws.cell(row=order_ws_rows_counter,
                              column=8).value = 'כן' if product['prining'] else 'לא'
                order_ws.cell(row=order_ws_rows_counter,
                              column=8).fill = header_fill
                order_ws.cell(row=order_ws_rows_counter,
                              column=8).alignment = align_rtl
                order_ws.cell(row=order_ws_rows_counter,
                              column=8).font = header_font
                order_ws.cell(row=order_ws_rows_counter,
                              column=8).border = bottom_border

                order_ws.cell(row=order_ws_rows_counter,
                              column=9).value = product['priningComment'] if product['prining'] else ''
                order_ws.cell(row=order_ws_rows_counter,
                              column=9).fill = header_fill
                order_ws.cell(row=order_ws_rows_counter,
                              column=9).alignment = align_rtl
                order_ws.cell(row=order_ws_rows_counter,
                              column=9).font = header_font
                order_ws.cell(row=order_ws_rows_counter,
                              column=9).border = bottom_border

                order_ws.cell(row=order_ws_rows_counter,
                              column=10).value = 'כן' if product['embroidery'] else 'לא'
                order_ws.cell(row=order_ws_rows_counter,
                              column=10).fill = header_fill
                order_ws.cell(row=order_ws_rows_counter,
                              column=10).alignment = align_rtl
                order_ws.cell(row=order_ws_rows_counter,
                              column=10).font = header_font
                order_ws.cell(row=order_ws_rows_counter,
                              column=10).border = bottom_border

                order_ws.cell(row=order_ws_rows_counter,
                              column=11).value = product['embroideryComment'] if product['embroidery'] else ''
                order_ws.cell(row=order_ws_rows_counter,
                              column=11).fill = header_fill
                order_ws.cell(row=order_ws_rows_counter,
                              column=11).alignment = align_rtl
                order_ws.cell(row=order_ws_rows_counter,
                              column=11).font = header_font
                order_ws.cell(row=order_ws_rows_counter,
                              column=11).border = bottom_border

                order_ws_rows_counter += 1

                entries = product['entries']
                # entry = ('color_name', 'size_name', 'varient_name', color_order, size_order): quantity
                # sort the entries by color name, size name, varient name based on the order of the colors and sizes
                sorted_entries = sorted(
                    entries.items(), key=lambda x: (x[0][3], x[0][4]))
                if len(sorted_entries) == 1 and sorted_entries[0][0][0].lower() == 'no color' and sorted_entries[0][0][1].lower() == 'one size':
                    continue
                sorted_entries = dict(sorted_entries)

                for entry in sorted_entries:  # product['entries']:
                    # entry = ('color_name', 'size_name', 'varient_name'): quantity
                    color = entry[0]
                    size = entry[1]
                    varient = entry[2]
                    quantity = product['entries'][entry]
                    order_ws.cell(row=order_ws_rows_counter,
                                  column=1).value = color
                    order_ws.cell(row=order_ws_rows_counter,
                                  column=1).alignment = align_rtl
                    # order_ws.cell(row=order_ws_rows_counter, column=1).font = order_ws.cell(row=order_ws_rows_counter, column=1).font.copy(bold=True)
                    order_ws.cell(row=order_ws_rows_counter,
                                  column=2).value = size
                    order_ws.cell(row=order_ws_rows_counter,
                                  column=2).alignment = align_rtl
                    # order_ws.cell(row=order_ws_rows_counter, column=2).font = order_ws.cell(row=order_ws_rows_counter, column=2).font.copy(bold=True)
                    order_ws.cell(row=order_ws_rows_counter,
                                  column=3).value = varient
                    order_ws.cell(row=order_ws_rows_counter,
                                  column=3).alignment = align_rtl
                    # order_ws.cell(row=order_ws_rows_counter, column=3).font = order_ws.cell(row=order_ws_rows_counter, column=3).font.copy(bold=True)
                    order_ws.cell(row=order_ws_rows_counter,
                                  column=4).value = quantity
                    order_ws.cell(row=order_ws_rows_counter,
                                  column=4).alignment = align_rtl
                    # order_ws.cell(row=order_ws_rows_counter, column=4).font = order_ws.cell(row=order_ws_rows_counter, column=4).font.copy(bold=True)
                    order_ws_rows_counter += 1

            # for order_product in order_products:
            #     product_name = order_product['title']

        headers = ['ברקוד', 'פריט', 'כמות כוללת',
                   'הערות', 'פירוט', 'ספק', 'כמות חדשה לספקים', 'האם שורה ראשית']
        # Add headers in bold
        for i in range(len(headers)):
            main_ws.cell(row=ws_rows_counter, column=i+1).value = headers[i]
            main_ws.cell(row=ws_rows_counter, column=i+1).font = main_ws.cell(
                row=ws_rows_counter, column=i+1).font.copy(bold=True)
        ws_rows_counter += 1
        for product_name in all_products:

            product = all_products[product_name]
            main_ws.cell(row=ws_rows_counter,
                         column=1).value = product['barcode']

            # main_ws.cell(row=ws_rows_counter, column=1).font = main_ws.cell(row=ws_rows_counter, column=1).font.copy(bold=True)
            main_ws.cell(row=ws_rows_counter, column=1).fill = header_fill
            main_ws.cell(row=ws_rows_counter, column=1).alignment = align_rtl
            main_ws.cell(row=ws_rows_counter, column=1).font = header_font
            main_ws.cell(row=ws_rows_counter, column=1).border = bottom_border

            main_ws.cell(row=ws_rows_counter, column=2).value = product_name
            main_ws.cell(row=ws_rows_counter, column=2).fill = header_fill
            main_ws.cell(row=ws_rows_counter, column=2).alignment = align_rtl
            main_ws.cell(row=ws_rows_counter, column=2).font = header_font
            main_ws.cell(row=ws_rows_counter, column=2).border = bottom_border

            main_ws.cell(row=ws_rows_counter,
                         column=3).value = product['total_quantity']
            main_ws.cell(row=ws_rows_counter, column=3).fill = header_fill
            main_ws.cell(row=ws_rows_counter, column=3).alignment = align_rtl
            main_ws.cell(row=ws_rows_counter, column=3).font = header_font
            main_ws.cell(row=ws_rows_counter, column=3).border = bottom_border

            main_ws.cell(row=ws_rows_counter,
                         column=4).value = product['comment']
            main_ws.cell(row=ws_rows_counter, column=4).fill = header_fill
            main_ws.cell(row=ws_rows_counter, column=4).alignment = align_rtl
            main_ws.cell(row=ws_rows_counter, column=4).font = header_font
            main_ws.cell(row=ws_rows_counter, column=4).border = bottom_border

            main_ws.cell(row=ws_rows_counter,
                         column=5).value = '\n'.join(product['details'])
            main_ws.cell(row=ws_rows_counter, column=5).fill = header_fill
            main_ws.cell(row=ws_rows_counter, column=5).alignment = align_rtl
            main_ws.cell(row=ws_rows_counter, column=5).font = header_font
            main_ws.cell(row=ws_rows_counter, column=5).border = bottom_border

            main_ws.cell(row=ws_rows_counter,
                         column=8).value = 'כן'
            main_ws.cell(row=ws_rows_counter, column=8).fill = header_fill
            main_ws.cell(row=ws_rows_counter, column=8).alignment = align_rtl
            main_ws.cell(row=ws_rows_counter, column=8).font = header_font
            main_ws.cell(row=ws_rows_counter, column=8).border = bottom_border

            # ws_rows_counter += 1
            # main_ws.cell(row=ws_rows_counter, column=1).value = 'צבע'
            # main_ws.cell(row=ws_rows_counter, column=1).font = main_ws.cell(row=ws_rows_counter, column=1).font.copy(bold=True)
            # main_ws.cell(row=ws_rows_counter, column=1).alignment = align_rtl
            # main_ws.cell(row=ws_rows_counter, column=2).value = 'מידה'
            # main_ws.cell(row=ws_rows_counter, column=2).font = main_ws.cell(row=ws_rows_counter, column=2).font.copy(bold=True)
            # main_ws.cell(row=ws_rows_counter, column=2).alignment = align_rtl
            # main_ws.cell(row=ws_rows_counter, column=3).value = 'ווריאנט'
            # main_ws.cell(row=ws_rows_counter, column=3).font = main_ws.cell(row=ws_rows_counter, column=3).font.copy(bold=True)
            # main_ws.cell(row=ws_rows_counter, column=3).alignment = align_rtl
            # main_ws.cell(row=ws_rows_counter, column=4).value = 'כמות'
            # main_ws.cell(row=ws_rows_counter, column=4).font = main_ws.cell(row=ws_rows_counter, column=4).font.copy(bold=True)
            # main_ws.cell(row=ws_rows_counter, column=4).alignment = align_rtl
            ws_rows_counter += 1
            entries = product['entries']
            # entry = ('color_name', 'size_name', 'varient_name', color_order, size_order): quantity
            # sort the entries by color name, size name, varient name based on the order of the colors and sizes
            sorted_entries = sorted(
                entries.items(), key=lambda x: (x[0][3], x[0][4]))
            if len(sorted_entries) == 1 and sorted_entries[0][0][0].lower() == 'no color' and sorted_entries[0][0][1].lower() == 'one size':
                continue
            sorted_entries = dict(sorted_entries)

            for entry in sorted_entries:  # product['entries']:
                # entry = ('color_name', 'size_name', 'varient_name'): quantity
                color = entry[0]
                size = entry[1]
                varient = entry[2]
                quantity = product['entries'][entry]
                main_ws.cell(row=ws_rows_counter, column=1).value = color
                main_ws.cell(row=ws_rows_counter,
                             column=1).alignment = align_rtl
                # main_ws.cell(row=ws_rows_counter, column=1).font = main_ws.cell(row=ws_rows_counter, column=1).font.copy(bold=True)
                main_ws.cell(row=ws_rows_counter, column=2).value = size
                main_ws.cell(row=ws_rows_counter,
                             column=2).alignment = align_rtl
                # main_ws.cell(row=ws_rows_counter, column=2).font = main_ws.cell(row=ws_rows_counter, column=2).font.copy(bold=True)
                main_ws.cell(row=ws_rows_counter, column=3).value = varient
                main_ws.cell(row=ws_rows_counter,
                             column=3).alignment = align_rtl
                # main_ws.cell(row=ws_rows_counter, column=3).font = main_ws.cell(row=ws_rows_counter, column=3).font.copy(bold=True)
                main_ws.cell(row=ws_rows_counter, column=4).value = quantity
                main_ws.cell(row=ws_rows_counter,
                             column=4).alignment = align_rtl
                # main_ws.cell(row=ws_rows_counter, column=4).font = main_ws.cell(row=ws_rows_counter, column=4).font.copy(bold=True)
                ws_rows_counter += 1

        # save file

        # date = datetime.datetime.now().strftime('%Y-%m-%d')
        ids = '_'.join([str(o['id']) for o in orders])
        filename = f'orders_{ids}.xlsx'
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(
            filename)
        wb.save(response)
        return response
    '''
    def export_to_excel(self, request, queryset):
        filesbuffers = []
        for morder in queryset:
            filesbuffers.append(morder.export_to_excel())
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
            for d in filesbuffers:
                file_name = d['name'] + '.xls'
                data = d['data']
                zip_file.writestr(file_name , data.getvalue())
        zip_buffer.seek(0)
        response = HttpResponse(
            zip_buffer.read(), content_type="application/zip")
        response['Content-Disposition'] = 'attachment; filename="products.zip"'
        return response
    '''


admin.site.register(MOrder, MOrderAdmin)


class MorderStatusAdmin(OrderedModelAdmin):
    list_display = ('name', 'move_up_down_links')
    # list_filter = ('name',)


admin.site.register(MorderStatus, MorderStatusAdmin)
