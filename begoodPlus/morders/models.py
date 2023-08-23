from colorfield.fields import ColorField
import re
from begoodPlus.secrects import FULL_DOMAIN
import pytz
from core.gspred import get_gspread_client
from gspread.cell import Cell
from gspread_formatting import *
from openpyxl.styles import Alignment
import copy

import gspread
from ordered_model.models import OrderedModelBase
from begoodPlus.secrects import SECRECT_BASE_MY_DOMAIN, SECRECT_CLIENT_SIDE_DOMAIN, ALL_MORDER_FILE_SPREEDSHEET_URL, ALL_PRICE_PROPOSAL_SPREEADSHEET_URL, ARCHIVED_MORDER_FILE_SPREEDSHEET_URL
from django.conf import settings
import reversion
from decimal import Decimal
import secrets
from django.forms import ValidationError
from django.urls import reverse
from django.utils.translation import gettext_lazy as _

from multiprocessing.connection import Client
import pandas as pd
from django.db import models
from django.contrib.auth.models import User
from core.utils import number_to_spreedsheet_letter
from catalogImages.models import CatalogImage
from client.models import Client
from color.models import Color
from core.models import SvelteCartModal
from productSize.models import ProductSize
from catalogImages.models import CatalogImageVarient
from provider.models import Provider
from django.utils.html import mark_safe
import datetime
from django.db.models import Count, F, Value
from django.db.models import OuterRef, Subquery
from django.db.models import Q
from django.contrib.postgres.aggregates import ArrayAgg
from django.db.models import Sum, Avg, When, Case
from django.db.models.functions import Substr
from django.db.models.functions import Concat
from django.db.models.functions import Length
from begoodPlus.secrects import SMARTBEE_DOMAIN, SMARTBEE_providerUserToken
import requests
from django.db.models.signals import pre_save, post_save, m2m_changed
from django.dispatch import receiver

# toOrder = models.IntegerField(default=0)


class MOrderItemEntry(models.Model):
    quantity = models.IntegerField(default=1)
    color = models.ForeignKey(
        to=Color, on_delete=models.SET_DEFAULT, default=76, null=True, blank=True)
    size = models.ForeignKey(
        to=ProductSize, on_delete=models.SET_DEFAULT, default=108, null=True, blank=True)
    varient = models.ForeignKey(
        to=CatalogImageVarient, on_delete=models.CASCADE, null=True, blank=True)
    # sheets_taken_quantity = models.IntegerField(default=0)
    # sheets_taken_quantity = models.CharField(
    # max_length = 255, null = True, blank = True)
    # sheets_provider = models.CharField(max_length=255, null=True, blank=True)

    def __str__(self):
        return str(self.quantity) + ' ' + str(self.color) + ' ' + str(self.size) + ' ' + str(self.varient)
    pass

    def validate_unique(self, *args, **kwargs):
        super().validate_unique(*args, **kwargs)
        if MOrderItemEntry.objects.filter(orderItem=self.orderItem.all().first(), color=self.color, size=self.size, varient=self.varient).count() > 1:
            raise ValidationError("This product has already been added")


class MOrderItem(models.Model):
    """
    This is the model for the items in the order.
    """
    product = models.ForeignKey(to=CatalogImage, on_delete=models.CASCADE)
    price = models.DecimalField(max_digits=10, decimal_places=2)
    # quantity = models.IntegerField(default=1)
    # color = models.ForeignKey(to=Color, on_delete=models.SET_DEFAULT,default=76, null=True, blank=True)
    # size = models.ForeignKey(to=ProductSize, on_delete=models.SET_DEFAULT, default=108, null=True, blank=True)
    # varient = models.ForeignKey(to=CatalogImageVarient, on_delete=models.CASCADE, null=True, blank=True)
    # provider = models.ForeignKey(to=Provider, on_delete=models.SET_DEFAULT, default=7)
    providers = models.ManyToManyField(to=Provider, blank=True)
    # clientProvider = models.CharField(max_length=255, null=True, blank=True)
    # clientBuyPrice = models.DecimalField(max_digits=10, decimal_places=2, null=True, blank=True)
    # ergent=models.BooleanField(default = False)
    # prining=models.BooleanField(default = False)
    # priningComment=models.CharField(
    #     max_length = 255, null = True, blank = True)
    # embroidery=models.BooleanField(default = False)
    # embroideryComment=models.CharField(
    #     max_length = 255, null = True, blank = True)
    comment = models.TextField(null=True, blank=True)
    entries = models.ManyToManyField(
        to=MOrderItemEntry, blank=True, related_name='orderItem')
    # taken = models.ManyToManyField(
    #     to=TakenInventory, blank=True, related_name='orderItem')
    # toProviders = models.ManyToManyField(
    #     to=ProviderRequest, blank=True, related_name='orderItem')
    prop_totalEntriesQuantity = property(lambda self: sum(
        [entry.quantity for entry in self.entries.all()]))
    prop_totalPrice = property(
        lambda self: self.prop_totalEntriesQuantity * self.price)
    created_at = models.DateTimeField(auto_now_add=True)

    class Meta:
        ordering = ['created_at', 'product__title', ]

    def __str__(self):
        # str(self.color) + " " + str(self.size) + (" " + self.varient.name) if self.varient != None else ' ' + str(self.quantity) + " " + str(self.price) + '₪'
        return str(self.product) + " | " + str(self.price) + '₪'
# Create your models here.


STATUS_CHOICES = [('new', 'חדש'), ('price_proposal', 'הצעת מחיר'), ('in_progress', 'סחורה הוזמנה'), ('in_progress2', 'מוכן לליקוט',), (
    'in_progress3', 'בהדפסה',), ('in_progress4', 'מוכן בבית דפוס'), ('in_progress5', 'ארוז מוכן למשלוח'), ('done', 'סופק'), ]

# LOCKED_CELL_COLOR = {
#     'backgroundColor': {'red': 0.9, 'green': 0.9, 'blue': 0.9},
#     'textFormat': {'fontSize': 12},
# }

BLANK_CELL_COLOR = {
    'backgroundColor': {'red': 1, 'green': 1, 'blue': 1},
    'textFormat': {'fontSize': 12},
    'borders': {
        'top': {
            'style': 'SOLID',
            'width': 1,
            'color': {'red': 0.9, 'green': 0.9, 'blue': 0.9},
        },
        'bottom': {
            'style': 'SOLID',
            'width': 1,
            'color': {'red': 0.9, 'green': 0.9, 'blue': 0.9},
        },
    },


}
PINK_HEADER_FORMAT_CONST = 'PINK_FORMAT_HEADER'
SUBTABLE_HEADER_FORMAT_CONST = 'SUBTABLE_HEADER_FORMAT'
USER_INPUT_FORMAT_CONST = 'USER_INPUT_FORMAT'
LOCKED_CELL_COLOR_CONST = 'LOCKED_CELL_COLOR'


@ reversion.register()
class MOrder(models.Model):
    created = models.DateTimeField(auto_now_add=True)
    updated = models.DateTimeField(auto_now=True)
    cart = models.ForeignKey(
        to=SvelteCartModal, on_delete=models.SET_NULL, null=True)
    client = models.ForeignKey(
        to=Client, on_delete=models.SET_NULL, null=True)
    agent = models.ForeignKey(
        to=User, on_delete=models.SET_NULL, null=True)
    name = models.CharField(max_length=100, blank=True,
                            null=True, default='')
    phone = models.CharField(max_length=100, blank=True,
                             null=True, default='')
    email = models.CharField(max_length=100, blank=True,
                             null=True, default='')
    status2 = models.ForeignKey(
        to='MorderStatus', on_delete=models.SET_NULL, null=True, blank=True)
    status_msg = models.TextField(_('status message'), blank=True, null=True)
    # order_type = models.CharField(max_length=100, choices=[('Invoice', 'חשבונית מס'), ('RefundInvoice', 'חשבונית זיכוי'), (
    #     'PriceProposal', 'הצעת מחיר'), ('ShippingCertificate', 'תעודת משלוח'), ('ReturnCertificate', 'תעודת החזרה')], blank=True, null=True)
    products = models.ManyToManyField(
        to=MOrderItem, blank=True, related_name='morder')
    message = models.TextField(null=True, blank=True)
    # freezeTakenInventory = models.BooleanField(default=False)
    # archive = models.BooleanField(default=False)
    # isOrder = models.BooleanField(default=False)
    # sendProviders = models.BooleanField(default=False)
    # startCollecting = models.BooleanField(default=False)
    prop_totalPrice = property(lambda self: sum(
        [item.prop_totalPrice for item in self.products.all()]))
    prop_totalPricePlusTax = property(
        lambda self: self.prop_totalPrice * Decimal('1.17'))

    total_sell_price = models.FloatField(
        _('total sell price'), default=0)
    last_status_updated = models.CharField(
        _('last status updated'), max_length=100, blank=True, null=True)
    last_notify_order_status = models.CharField(
        max_length=100, blank=True, null=True)
    last_notify_order_total_price = models.FloatField(
        _('last notify order total price'), default=0)
    # save

    def recalculate_total_price(self):
        if self.pk:
            # new_price = self.prop_totalPrice
            # if self.total_sell_price != new_price:
            #     self.total_sell_price = new_price
            # pass

            # we iterete over all products
            # for each product we sum the total amount and multiply it by the price
            # we sum all the products total price
            # we compare the total price to self.total_sell_price
            total_price = 0
            for product in self.products.all():
                # product.price
                # sum(product.entries.all().quantity)
                total_price += product.price * \
                    sum(product.entries.all().values_list('quantity', flat=True))
            if self.total_sell_price != total_price:
                self.total_sell_price = total_price
            return total_price
        else:
            return 0

    def notify_order_status_update(self):
        from morders.tasks import send_morder_status_update_to_telegram
        total_sell = self.recalculate_total_price()
        print('notify_order_status_update_post_save: ', total_sell)

        if total_sell > 0:
            # we compare self.total_sell_price to self
            # kkkkkkkkk

            status = self.get_status_display()

            # we check if last_notify_order_status is not the same as status
            # or last_notify_order_total_price is not the same as total_sell_price
            if self.last_notify_order_status != status.name or self.last_notify_order_total_price != total_sell:
                if settings.DEBUG:
                    send_morder_status_update_to_telegram(morder_id=self.id)

                else:
                    send_morder_status_update_to_telegram.delay(
                        morder_id=self.id)
                    pass
                self.last_notify_order_status = status.name
                self.last_notify_order_total_price = total_sell
                self.save()
            else:
                print('no need to send notification')
                pass
        print('done notify_order_status_update_post_save: ', total_sell)
        pass

    def update_sell_price_from_price_proposal_sheet(self, sheet_id=None):
        #
        if not sheet_id:
            sheet_id = self.price_proposal_sheetid
        if not sheet_id:
            return ['sheet_id is None']
        gspred_client = get_gspread_client()
        workbook = gspred_client.open_by_url(
            ALL_PRICE_PROPOSAL_SPREEADSHEET_URL)

        ws = list(filter(lambda ws: ws.id == int(
            sheet_id), workbook.worksheets()))

        if not ws:
            return ['sheet_id ' + sheet_id + ' not found in ' + str(workbook.worksheets())]
        ws = ws[0]
        ROW_OFFSET = 5
        PRODUCT_NAME_COL = 1
        SELL_PRICE_COL = 8
        erorrs = []
        for row in range(ROW_OFFSET, ws.row_count):
            product_name = ws.cell(row, PRODUCT_NAME_COL).value
            sell_price = ws.cell(row, SELL_PRICE_COL).value
            if not product_name:
                break
            try:
                product = self.products.get(product__title=product_name)
                product.price = sell_price
                product.save()
            except Exception as e:
                print(e)
                erorrs.append(
                    f'error in  {product_name} \n' + str(e))
        return erorrs

    def get_sheets_order_link(self):
        ret = ''
        if self.gid:
            if self.order_sheet_archived:
                ret = ARCHIVED_MORDER_FILE_SPREEDSHEET_URL
            else:
                ret = ALL_MORDER_FILE_SPREEDSHEET_URL
            ret = re.sub(r'#gid=\d+', f'#gid={self.gid}',
                         ret)
        return ret

    def get_sheets_price_prop_link(self):
        ret = ''
        if self.price_proposal_sheetid:
            ret = ALL_PRICE_PROPOSAL_SPREEADSHEET_URL
            ret = re.sub(r'#gid=\d+', f'#gid={self.price_proposal_sheetid}',
                         ALL_PRICE_PROPOSAL_SPREEADSHEET_URL)

        return ret

    def save(self, *args, **kwargs):
        from docsSignature.utils import create_signature_doc_from_morder
        try:
            # create_signature_doc_from_morder(self)
            pass
        except Exception as e:
            print(e)

        super().save(*args, **kwargs)

    class Meta:
        ordering = ['-created']

    def get_exel_data(self):
        # שם	כמות	מחיר מכירה ללא מע"מ	ספקים
        products = []
        qs = self.products.all().select_related('product').prefetch_related(
            'entries', 'providers', 'entries__color', 'entries__size', 'entries__varient')
        for item in qs:
            # Entry: quantity,color,size, varient
            entries = {}
            defualt_color = Color.objects.get(pk=76)
            defualt_size = ProductSize.objects.get(pk=108)
            for entry in item.entries.all():
                color = entry.color.name if entry.color != None else defualt_color.name
                color_order = entry.color.name if entry.color != None else defualt_color.name
                size = entry.size.size if entry.size != None else defualt_size.size
                size_order = entry.size.code if entry.size != None else defualt_size.code
                varient = entry.varient.name if entry.varient != None else ''
                # provider = entry.provider if entry.provider != None else ''
                # taken = entry.sheets_taken_quantity if entry.sheets_taken_quantity != None else ''
                # provider = entry.sheets_provider if entry.sheets_provider != None else ''
                key = tuple([color, size, varient, color_order, size_order])
                # entries.append([color, size,varient,entry.quantity])
                entries[key] = entry.quantity
            item_data = {'title': item.product.title, 'total_quantity': item.prop_totalEntriesQuantity,
                         'price': item.price,
                         'price_tax': item.price * Decimal('1.17'),
                         'providers': ','.join([provider.name for provider in item.providers.all()]),
                         'comment': item.comment if item.comment != None else '',
                         'barcode': item.product.barcode if item.product.barcode != None else '',
                         'entries': entries}
            products.append(item_data)
        data = {
            'name': self.name if self.name != None else self.client.business_name if self.client.business_name != None else '',
            'products': products,
            'message': self.message if self.message != None else '',
            'date': self.created.strftime('%Y_%m_%d'),
            'id': self.id,
            'status': self.status2.name if self.status2 != None else '',
            'admin_link': self.get_edit_url_without_html(base_url=FULL_DOMAIN),
        }
        return data

    def view_morder_pdf_link(self):
        link = reverse('view_morder_pdf', args=(self.pk,))
        return mark_safe('<a href="{}">{}</a>'.format(link, 'הצג הזמנה'))

    def get_edit_url_without_html(self, base_url=None):
        link = reverse('admin_edit_order', args=(self.pk,))
        if base_url:
            return base_url + link
        return link

    def get_edit_url_id(self, base_url=None):
        link = reverse('admin_edit_order', args=(self.pk,))
        if base_url:
            return base_url + link
        return mark_safe('<a href="{}">{}</a>'.format(link, self.pk))
    get_edit_url_id.short_description = 'ID'

    def products_display(self):
        products = []
        qs = self.products.all().prefetch_related(
            'product', 'size', 'color', 'varient', 'provider')
        for p in qs:
            size = p.size.size if p.size != None else ' '
            color = p.color.name if p.color != None else ' '
            verient = p.varient.name if p.varient != None else ' '
            products.append({'product': p.product.title, 'quantity': int(p.quantity), 'price': p.price, 'color': color, 'size': size, 'varient': verient,
                            'comment': p.comment, 'clientBuyPrice': p.clientBuyPrice, 'clientProvider': p.clientProvider, 'provider': p.provider.name})
        try:
            df = pd.DataFrame(products)
        except Exception as e:
            print(e)
            return ' '
        print(df.columns)
        print(df.head())
        # df['cell_display'] = df['quantity'].astype(str)# + ' ' + df['price'].astype(str) + '₪'
        # df['price_display'] = df['price'] + '₪'
        if len(df) == 0:
            return mark_safe(df.to_html(index=False))
        df = df.pivot(index=['product', 'color', 'varient',
                             'price'], columns='size', values=['quantity'])

        print(df.columns)
        print(df.head())
        html = df.to_html(index=True, header=True,
                          table_id='table_id', na_rep='-')
        return mark_safe(html)

    def get_status_display(self):
        status = self.status2
        return status
        # STATUS_CHOICES = [('new', 'חדש'), ('price_proposal',....
        # for choice in STATUS_CHOICES:
        #     if choice[0] == status:
        #         return choice[1]

    def get_edit_order_url(self):
        return reverse('admin_edit_order', args=(self.pk,))
        # morders/edit-order/{self.pk}/
# @receiver(pre_save, sender=MOrder)
# def check_for_status_update(sender, instance, *args, **kwargs):

#     pass


class MorderStatus(OrderedModelBase):
    COLOR_PALETTE = [
        ('#FFFFFF', 'white', ),
        ('#000000', 'black', ),
        ('#FFFFFF00', 'transparent', ),
    ]
    name = models.CharField(max_length=100, unique=True,
                            verbose_name=_('name'))
    sort_order = models.PositiveIntegerField(editable=False, db_index=True)
    color = ColorField(verbose_name=_(
        'color'), default='#FF12E2FF', format='hexa', samples=COLOR_PALETTE)
    order_field_name = "sort_order"

    class Meta:
        ordering = ("sort_order",)

    def __str__(self):
        return self.name
