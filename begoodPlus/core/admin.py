from django.utils.html import mark_safe
from django.utils.http import urlencode
import json
from core.models import UserProductPhoto
from .models import SvelteContactFormModal
from .models import BeseContactInformation
from calendar import c
import requests
from django.contrib import admin
from django.utils.translation import gettext_lazy as _
import io
from django.http import FileResponse
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import client
from openpyxl.drawing import image
from openpyxl.utils.cell import get_column_letter

from client.models import UserQuestion
# Register your models here.
from .models import ActiveCartTracker, SvelteCartModal, SvelteCartProductEntery, UserSearchData


class ActiveCartTrackerAdmin(admin.ModelAdmin):
    list_display = ('last_updated', 'created_at', 'last_ip',
                    'active_cart_id', 'cart_products_size')
    readonly_fields = ('last_updated', 'created_at', 'last_ip', 'active_cart_id',
                       'products_amount_display_with_sizes_and_colors', 'cart_products_size')
    ordering = ('-last_updated',)


admin.site.register(ActiveCartTracker, ActiveCartTrackerAdmin)


class UserSearchDataAdmin(admin.ModelAdmin):
    list_display = ('id', 'created_date', 'term', 'resultCount', 'session')


admin.site.register(UserSearchData, UserSearchDataAdmin)


class BeseContactInformationAdmin(admin.ModelAdmin):
    list_display = ('id', 'formUUID', 'url', 'created_date', 'name',
                    'email', 'phone', 'message', 'sumbited', 'owner_display')


admin.site.register(BeseContactInformation, BeseContactInformationAdmin)


class ContactFormModalAdmin(admin.ModelAdmin):
    list_display = ('user', 'uniqe_color', 'device', 'name',
                    'email', 'phone', 'message', 'created_date')


admin.site.register(SvelteContactFormModal, ContactFormModalAdmin)


class UserProductPhotoAdmin(admin.ModelAdmin):
    list_display = ('user', 'photo_display', 'created_date',
                    'buy_price', 'want_price', 'description',)
    readonly_fields = ('photo_display',)


admin.site.register(UserProductPhoto, UserProductPhotoAdmin)


class SvelteCartProductEnteryAdmin(admin.ModelAdmin):
    list_display = ('id', 'product', 'amount', 'details')


admin.site.register(SvelteCartProductEntery, SvelteCartProductEnteryAdmin)


class SvelteCartModalAdmin(admin.ModelAdmin):
    list_display = ('id', 'doneOrder', 'user', 'buiss_display', 'cart_count',
                    'uniqe_color', 'name', 'phone', 'email', 'created_date', 'agent', 'copy_cart')

    #filter_horizontal = ('products',)
    readonly_fields = ('buiss_display', 'productsRaw',
                       'cart_count', 'copy_cart')
    readonly_fields = ('productsRaw', 'copy_cart')
    exclude = ('products', 'productEntries')

    def copy_cart(self, obj):
        url_res = {}
        raw = json.loads(obj.productsRaw or '{}')
        try:
            for key in raw:
                url_res[key] = {'mentries': raw[key]
                                ['mentries'], 'amount': raw[key]['amount']}
            url = urlencode({'cart_json': json.dumps(url_res)})
        except Exception as e:
            url = ''
        ms_domain = 'https://www.ms-global.co.il/?'
        boost_domain = 'https://boost-pop.com/?'
        local_domain = 'http://127.0.0.1:3000/?'
        return mark_safe('<ul><li><a href="{}{}">{}</a></li><li><a href="{}{}">{}</a></li><li><a href="{}{}">{}</a></li></ul>'
                         .format(ms_domain, url, 'ms-global',
                                 boost_domain, url, 'boost-pop',
                                 local_domain, url, 'local'))

    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        extra_context['my_data'] = {'object_id': object_id}
        return super(SvelteCartModalAdmin, self).change_view(
            request, object_id, form_url, extra_context=extra_context,
        )

    def buiss_display(self, obj):
        if obj.user:
            if obj.user.is_anonymous == False:
                return obj.user.client.businessName
            else:
                return 'anonymous'
    buiss_display.short_description = _('Business Name')

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        qs = qs.select_related('user').select_related(
            'user__client').prefetch_related('productEntries')
        return qs
    readonly_fields = ('products_amount_display_with_sizes_and_colors',)
    actions = ['resend_email_action',
               'download_cart_excel', 'turn_to_morder', ]
    list_select_related = ['user']

    def resend_email_action(modeladmin, request, queryset):
        for obj in queryset:
            send_cart_notification.delay(obj.id)
    resend_email_action.short_description = _("resend selected carts email")

    def turn_to_morder(self, request, queryset):
        for obj in queryset:
            obj.doneOrder = True
            obj.turn_to_morder()
            obj.save()
    turn_to_morder.short_description = _("turn selected carts to order")

    def data_to_excel(self, data, filename):
        wb = Workbook()
        main_ws = wb.active
        main_ws.sheet_view.rightToLeft = True
        ws_rows_counter = 1
        main_ws.append(['שם מוצר', 'כמות', 'מחיר עלות ללא מע"מ', 'מחיר חנות ללא מע"מ',
                       'מחיר לקוח פרטי ללא מע"מ', 'ספקים', 'לקוח', 'תאריך הזמנה'])
        for cart in data:
            active_ws = wb.create_sheet(cart['משתמש'])
            active_ws.sheet_view.rightToLeft = True
            active_ws.append(['משתמש', 'מכשיר', 'שם בטופס',
                             'טלפון בטופס', 'אימייל בטופס', 'תאריך הגשה'])
            active_ws.append([cart['משתמש'], cart['מכשיר'], cart['שם בטופס'],
                             cart['טלפון בטופס'], cart['אימייל בטופס'], cart['תאריך הגשה']])
            active_ws.append(['מוצרים'])
            active_ws.append(['שם', 'כמות', 'מחיר עלות ללא מע"מ',
                             'מחיר חנות ללא מע"מ', 'מחיר לקוח פרטי ללא מע"מ', 'ספקים'])

            row_products_offset = 4
            current_row = row_products_offset
            for product in cart['מוצרים']:

                active_ws.append([product['שם'], product['כמות'], product['מחיר עלות ללא מע"מ'],
                                 product['מחיר חנות ללא מע"מ'], product['מחיר לקוח פרטי ללא מע"מ']])
                providers = product['ספקים']
                if(len(providers) > 0):
                    providers = [(v['name'] + ' - ' + v['makat']).replace(',',
                                                                          '').replace('\"', '') for v in providers]
                    print(providers)
                    providers_str = ','.join(providers)
                    dv = DataValidation(
                        type="list", formula1='"' + providers_str + '"', allow_blank=False)
                    active_ws.add_data_validation(dv)
                    current_row += 1
                    wanted_col = 6
                    provider_cell = active_ws.cell(
                        row=current_row, column=wanted_col)
                    provider_cell.value = providers[0]
                    dv.add(provider_cell)

                    ws_rows_counter += 1
                    main_ws.append([product['שם'], product['כמות'], product['מחיר עלות ללא מע"מ'],
                                   product['מחיר חנות ללא מע"מ'], product['מחיר לקוח פרטי ללא מע"מ']])
                    main_ws.add_data_validation(dv)
                    provider_cell = main_ws.cell(
                        row=ws_rows_counter, column=wanted_col)
                    provider_cell.value = providers[0]
                    dv.add(provider_cell)
                    main_ws.cell(row=ws_rows_counter,
                                 column=7).value = cart['משתמש']
                    main_ws.cell(row=ws_rows_counter,
                                 column=8).value = cart['תאריך הגשה']
        return wb


admin.site.register(SvelteCartModal, SvelteCartModalAdmin)


class UserQuestionAdmin(admin.ModelAdmin):
    list_display = ('id', 'is_answered', 'user', 'product',
                    'question', 'answer', 'created_at', 'ip',)
    search_fields = ('user', 'question', 'answer', 'product')


admin.site.register(UserQuestion, UserQuestionAdmin)
